import traceback
from difflib import SequenceMatcher

from lxml import html

import requests
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException

import sqlite3

from tweepstr import get_tweets

conn = sqlite3.connect('companies2.db')
cur = conn.cursor()
# cur.execute('CREATE TABLE companies (project_number text, client_id text, input_name text, title text, type text, industry text, founded text, founder text, headquarters text, key_people text, products text, revenue text, operating_income text, net_income text, owners text, website text, traded_as text, subsidiaries text, area_served text, parent text, divisions text, link text, tweets text)')
query = 'INSERT INTO companies VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'

# print(requests.__version__)

wb = load_workbook('client_list.xlsx')
ws = wb.active
def get_info(tree, project_no, client_id, input_name, url):
    info = tree.xpath('//*[@id="mw-content-text"]/div/table')[0]
    title = tree.xpath('//*[@id="firstHeading"]')[0].text_content()
    type = ''
    industry = ''
    founded = ''
    founder = ''
    headquarters = ''
    key_people = ''
    products = ''
    revenue = ''
    operating_income = ''
    net_income = ''
    owners = ''
    website = ''

    traded_as = ''
    subsidiaries = ''

    area_served = ''
    parent = ''
    divisions = ''
    for row in info:
        try:
            thead = row.xpath('.//th')[0]
            field = row.xpath('.//td')[0]
            t = thead.text_content().lower()
            f = field.text_content().strip()
            if 'traded' in info.text_content().lower():
                traded_as = field.text_content().strip()
            if 'type' in thead.text_content().lower():
                type = field.text_content().strip()
            if 'industry' in thead.text_content().lower():
                industry = field.text_content().strip()
            if 'founded' in thead.text_content().lower():
                founded = field.text_content().strip()
            if 'founder' in thead.text_content().lower():
                founder = field.text_content().strip()
            if 'headquarters' in thead.text_content().lower():
                headquarters = field.text_content().strip()
            if 'key people' in thead.text_content().lower():
                key_people = field.text_content().strip()
            if 'products' in thead.text_content().lower():
                products = field.text_content().strip()
            if 'revenue' in thead.text_content().lower():
                revenue = field.text_content().strip()
            if 'operating income' in thead.text_content().lower():
                operating_income = field.text_content().strip()
            if 'net income' in thead.text_content().lower():
                net_income = field.text_content().strip()
            if 'owne' in thead.text_content().lower():
                owners = field.text_content().strip()
            if 'website' in thead.text_content().lower():
                website = field.text_content().strip()
            if 'traded as' in thead.text_content().lower():
                traded_as = field.text_content().strip()
            if 'subsidiaries' in thead.text_content().lower():
                subsidiaries = field.text_content().strip()

            if 'area' in thead.text_content().lower():
                area_served = field.text_content().strip()
            if 'parent' in thead.text_content().lower():
                parent = field.text_content().strip()
            if 'division' in thead.text_content().lower():
                divisions = field.text_content().strip()
        except:
            # traceback.print_exc()
            continue
    tweets = ''.join(get_tweets(input_name))
    cur.execute(query, (project_no, client_id, input_name, title, type, industry, founded, founder, headquarters, key_people, products, revenue, operating_income, net_income, owners, website, traded_as, subsidiaries, area_served, parent, divisions, url, tweets))
    conn.commit()
    # print('\nd\n')

def _gen_tree(url):
    page = requests.get(url)
    tree = html.fromstring(page.content)
    page.close()
    return tree, page.url

client_ids = [tuple[0] for tuple in cur.execute('select client_id from companies')]
base_url = 'https://en.wikipedia.org'
for i in range(2, 7501):
    try:
        project_no = ws['A' + str(i)].value
        client_id = ws['B' + str(i)].value
        if client_id in client_ids:
            # print(i)
            continue
        name = ws['C' + str(i)].value
        print('Processing ' + name)
        url = 'https://en.wikipedia.org/w/index.php?search=' + name.replace(' ', '+').replace('&', '%26')
        tree, og_url = _gen_tree(url)
        # wait_for_xpath(driver, '//ul[@class="mw-search-results"]', time=3)
        results = tree.xpath('//ul[@class="mw-search-results"]/li')
        if len(results) < 1:
            try:
                get_info(tree, project_no, client_id, name, og_url)
            except:
                tweets = ''.join(get_tweets(name))
                cur.execute(query, (project_no, client_id, name, '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', tweets))
                conn.commit()
            continue
        matches = []
        for result in results:
            res_text = result.text_content()
            s = SequenceMatcher(None, name, result.text_content().split('   ')[0])
            # if s.ratio() > .65:
            #     print(name, '||||||||', result.text_content().split('   ')[0], s.ratio())
            if name in result.text_content().split('   ')[0] or result.text_content().split('   ')[0] in name or s.ratio() > .70:
                # print('MATCH!', name, '||||||||', result.text_content().split('   ')[0], s.ratio())
                matches.append(result.xpath('./div[1]/a')[0].get('href'))
            else:
                if name in res_text:
                    # print(name, result.text_content())
                    # print()
            # print(result.text)
        match_successes = 0
        for match in matches:
            try:
                # pass
                tree, og_url = _gen_tree(base_url + match)
                get_info(tree, project_no, client_id, name, og_url)
                match_successes += 1
            except:
                traceback.print_exc()
                pass
        if match_successes == 0 or len(matches) == 0:
            tweets = ''.join(get_tweets(name))
            cur.execute(query, (project_no, client_id, name, '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', tweets))
            conn.commit()
    except TimeoutException:
        traceback.print_exc()
        get_info(tree, project_no, client_id, name, og_url)
print('Done')
